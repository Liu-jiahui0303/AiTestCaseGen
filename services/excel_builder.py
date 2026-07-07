import io
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import EXCEL_HEADERS, EXCEL_COL_WIDTHS, PRIORITY_FILLS

logger = logging.getLogger(__name__)

# Sheet 名称非法字符 & 31 字符长度限制
_SHEET_NAME_RE = re.compile(r"[\\\*\[\]\/\?:]")
_SHEET_NAME_MAX = 31

# 样式（预创建，避免每行重复创建）
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_BODY_FONT = Font(name="微软雅黑", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_WRAP_ALIGN = Alignment(vertical="center", wrap_text=True)
_FIELD_KEYS = ["id", "module", "title", "precondition", "steps", "expected", "type", "priority"]


def _clean_sheet_name(name: str) -> str:
    cleaned = _SHEET_NAME_RE.sub("", name)
    return cleaned[:_SHEET_NAME_MAX]


def _write_sheet(ws, cases: list):
    """向已创建的工作表写入表头+数据+格式"""
    # 表头
    for col, title in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER

    # 数据行
    for row_idx, tc in enumerate(cases, 2):
        values = [tc.get(k, "") for k in _FIELD_KEYS]
        fc = PRIORITY_FILLS.get(tc.get("priority", ""))
        row_fill = PatternFill(start_color=fc[0], end_color=fc[1], fill_type="solid") if fc else None

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN if col_idx in (1, 7, 8) else _WRAP_ALIGN
            if row_fill:
                cell.fill = row_fill

    # 列宽 / 冻结 / 筛选
    for i, w in enumerate(EXCEL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{len(cases) + 1}"


def build_excel(test_cases: list) -> io.BytesIO:
    """根据测试用例列表生成 Excel，包含汇总 sheet + 每个模块一个 sheet"""
    wb = Workbook()

    # ── 汇总 sheet（全部用例） ──
    ws = wb.active
    ws.title = "测试用例"
    _write_sheet(ws, test_cases)

    # ── 按模块分组 ──
    modules: dict[str, list] = {}
    for tc in test_cases:
        m = tc.get("module", "未分类") or "未分类"
        modules.setdefault(m, []).append(tc)

    # ── 每个模块一个 sheet ──
    for mod_name in sorted(modules):
        sheet_name = _clean_sheet_name(mod_name)
        # 处理重名（Excel 不允许同名 sheet）
        base = sheet_name
        n = 2
        while sheet_name in {s.title for s in wb.worksheets}:
            suffix = f" ({n})"
            sheet_name = f"{base[:_SHEET_NAME_MAX - len(suffix)]}{suffix}"
            n += 1
        mws = wb.create_sheet(title=sheet_name)
        _write_sheet(mws, modules[mod_name])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Excel generated: %d cases, %d sheets", len(test_cases), len(wb.worksheets))
    return output
